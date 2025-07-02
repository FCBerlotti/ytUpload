import pyautogui as pg
import time
import pyperclip
import requests
import subprocess
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from conteudos import conteudos

consoleCounter = 0
videoUploaded = False
contador = 0
time_soma = 0
errorInfo = False
version = "1.2"

def ntfy(message):
    try:
        requests.post(('https://ntfy.sh/ytUploadNotifier'), data=(message).encode(encoding='utf-8'))
    except requests.exceptions.HTTPError as errh:
        print("Http Error:", errh)
    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting:", errc)
    except requests.exceptions.Timeout as errt:
        print("Timeout Error:", errt)
    except requests.exceptions.RequestException as err:
        print("OOps: Something Else", err)

def timeCalc():
    global tempo_estimado, horario_estimado, time_soma, contador, time_final, time_start, end, videoNumber
    time_final = time.time()
    
    contador += 1
    time_soma += time_final-time_start
    tempo_estimado = (time.time() + ((time_soma / contador)*(end-videoNumber)))
    horario_estimado = datetime.fromtimestamp((time.time() + ((time_soma / contador)*(end-videoNumber)))).strftime('%Y-%m-%d %H:%M:%S')

def openNav(site):
    caminho_chrome = "C:/Program Files/Google/Chrome/Application/chrome.exe"  # Atualize o caminho se necessário
    comando = [caminho_chrome, site]
    subprocess.Popen(comando)

def openConsole():
    global consoleCounter
    pg.hotkey("ctrl", "shift", "i")
    for i in range(0,100):
        inspecActivity = pg.locateCenterOnScreen("images/console/inspecActivity.png", confidence=0.95)
        if inspecActivity is not None and len(inspecActivity) == 2:
            pg.hotkey("ctrl", "p")
            break
        time.sleep(0.1)
    time.sleep(1)
    pg.hotkey('ctrl', 'a')
    time.sleep(0.5)
    pg.write(">console")
    time.sleep(0.5)
    panelVerify = pg.locateCenterOnScreen("images/console/panel.png", confidence=0.95)
    if panelVerify is not None and len(panelVerify) == 2:
        time.sleep(0.5)
        pg.press("enter")
        time.sleep(2)
        pg.write("allow pasting")
        time.sleep(0.5)
        pg.press("enter")
        time.sleep(0.5)
        pg.press('up')
        time.sleep(0.5)
        pg.press('enter')
    else:
        pg.hotkey("ctrl", "shift", "i")
        if consoleCounter < 2:
            openConsole()
            consoleCounter += 1

def execJs(path):
    try:
        # Abrir o arquivo JS e ler o conteúdo
        with open(path, "r", encoding="utf-8") as file:
            js_script = file.read()

            # Copiar o conteúdo do script para a área de transferência
            pyperclip.copy(js_script)

            # Usar pg para colar o script no console
            pg.hotkey('ctrl', 'v')  # Cola o conteúdo copiado
            time.sleep(0.25)
            pg.press('enter')  # Pressiona Enter para executar o script
            time.sleep(1)

    except FileNotFoundError:
        print(f"O arquivo {path} não foi encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def clickImage(click, imageSelect, conf, attempts):
    for i in range(0, attempts):
        imageFind = pg.locateCenterOnScreen(imageSelect, confidence=conf)
        if imageFind is not None and len(imageFind) == 2:
            if click is True:
                pg.moveTo(imageFind[0], imageFind[1])
                time.sleep(0.5)
                pg.click()
                time.sleep(0.5)

            return True
        time.sleep(0.1)
    print(f"IMAGEM NAO ENCONTRADA - {imageSelect}")
    return False

def errorFunction():
    global errorInfo, videoNumber, firstDate, jumpDay
    errorInfo = True
    firstDate += jumpDay
    closeNav()
    logs.salvar_dados_excel('Error')
    timeCalc()
    ntfy(f"❌❌❌❌❌\nPrevisão de termino: {horario_estimado}\nContas Verficadas {videoNumber}/{end}\nCategoria: {videoType}")

def closeNav():
    pg.hotkey('ctrl', 'w')
    time.sleep(1)
    pg.press('enter')
    time.sleep(1)
    pg.hotkey("alt", "f4")
    time.sleep(0.5)
    pg.press('esc')

def aboutVideoInfos():
    global foundSelector, tittle, desc, tittlePath

    try:
        with open(f"{tittlePath}", "r", encoding="utf-8") as file:
            tittle = file.read()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{tittlePath}' não foi encontrado.")
    except Exception as e:
        print(f"Erro ao ler o arquivo: {str(e)}")
    
    try:
        with open(f"{descPath}", "r", encoding="utf-8") as arquivo:
            desc = arquivo.read()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{tittlePath}' não foi encontrado.")
    except Exception as e:
        print(f"Erro ao ler o arquivo: {str(e)}")
           
def dateCalculate():
    global dateSelect, firstDate
    today = datetime.now()
    dateSelect = (today + timedelta(days=firstDate)).strftime("%d/%m/%Y")

class steps:
    def step1Upload():
        global videoNumber
        time.sleep(5)
        execJs('js/youtube/upload.js')
        time.sleep(1)
        pg.hotkey('ctrl', 'shift', 'i')
        time.sleep(2)
        clickImage(True, 'images/youtube/upload.png', 0.9, 50)
        time.sleep(0.5)
        linkFolderClick = clickImage(True, 'images/windows/linkFolder.png', 0.9, 20)
        if linkFolderClick is True:
            pg.write(foundSelector)
            time.sleep(1)
            pg.press('enter')
            time.sleep(1.5)
            pg.hotkey('ctrl', 'f')
            time.sleep(1)
            pg.write(f'"vid{videoNumber}.mp4"', 0.15)
            time.sleep(0.75)
            selectVideo = clickImage(True, 'images/windows/mp4.png', 0.9, 80)
            if selectVideo is True:
                time.sleep(0.5)
                pg.press('enter')
                time.sleep(0.5)
        else:
            errorFunction()
            return
    
    def step2Edit():
        global videoUploaded
        # Editar titulo e descrição
        clickImage(True, 'images/youtube/videoTittle.png', 0.9, 100)
        time.sleep(1)
        pg.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pg.write(tittle)
        time.sleep(5)
        clickImage(True, 'images/youtube/videoDesc.png', 0.88, 50)
        time.sleep(1)
        pg.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pyperclip.copy(desc)
        time.sleep(0.5)
        pg.hotkey('ctrl', 'v')
        time.sleep(10)

        # Selecionar capa do video
        openConsole()
        time.sleep(1)
        execJs('js/youtube/thumbnail.js')
        time.sleep(2)
        linkFolderClick = clickImage(True, 'images/windows/linkFolder.png', 0.9, 20)
        if linkFolderClick is True:
            pg.write(foundSelector)
            time.sleep(1)
            pg.press('enter')
            time.sleep(1.5)
            pg.hotkey('ctrl', 'f')
            time.sleep(1)
            pg.write(f'"thumbnail.png"')
            selectThumb = clickImage(True, 'images/windows/thumbnail.png', 0.9, 80)
            if selectThumb is True:
                time.sleep(0.5)
                pg.press('enter')
                time.sleep(0.5)
        time.sleep(5)
        pg.hotkey("ctrl", 'shift', "i")
        time.sleep(1)
        openConsole()
        time.sleep(1)
        execJs('js/youtube/next.js')
        time.sleep(1)

        copyright = clickImage(False, 'images/youtube/copyright.png', 0.85, 1)
        copyright2 = clickImage(False, 'images/youtube/copyright2.png', 0.85, 1)

        for i in range(0, 300):
            if copyright is False and copyright2 is False:
                copyright = clickImage(False, 'images/youtube/copyright.png', 0.85, 1)
                copyright2 = clickImage(False, 'images/youtube/copyright2.png', 0.85, 1)
            else:
                break
            time.sleep(0.1)

        if copyright is True or copyright2 is True:
            time.sleep(1)
            execJs('js/youtube/next2.js')
            time.sleep(1)
            verifyPage = clickImage(False, 'images/youtube/verifyPage.png', 0.9, 120)
            if verifyPage is True:
                execJs('js/youtube/date.js')
                time.sleep(5)
                dateChange = clickImage(True, 'images/youtube/2025.png', 0.9, 40)
                if dateChange is True:
                    time.sleep(1.5)
                    pg.hotkey('ctrl', 'a')
                    time.sleep(0.5)
                    dateCalculate()
                    pg.write(f'{dateSelect}') 
                    time.sleep(1)
                    pg.press('enter')
                    time.sleep(1)
                    hourSelect = clickImage(True, 'images/youtube/hourSelect.png', 0.9, 50)
                    time.sleep(0.75)
                    if hourSelect is True:
                        pg.hotkey('ctrl', 'a')
                        time.sleep(1)
                        pg.write(hour, 0.05)
                        time.sleep(0.5)
                        pg.press('enter')
                        time.sleep(1)
                    else:
                        errorFunction()
                        return
                    clickImage(True, 'images/youtube/programar.png', 0.9, 40)
                    time.sleep(1)
            videoPosted = clickImage(False, 'images/youtube/videoPosted.png', 0.88, 150)
            if videoPosted is True:
                videoUploaded = True
            else:
                videoUploaded = False
                errorFunction()
                return
        else:
            errorFunction()
            return
        
class logs:
    def colorir_celulas(arquivo_excel):
        try:
            # Carregar o arquivo Excel
            wb = openpyxl.load_workbook(arquivo_excel)
            sheet = wb.active

            # Definir cores para os textos
            verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            roxo = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
            branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Dicionário mapeando textos para cores
            cores = {
                "Error": roxo,
                "Posted": verde,
            }

            # Loop pelas células do intervalo especificado
            for linha in range(1, 2400 + 1): 
                for coluna in range(1, 3):  # Colunas de A até C(3)
                    celula = sheet.cell(row=linha, column=coluna)
                    # Verifica se o valor da célula existe e está no dicionário de cores
                    if celula.value and str(celula.value) in cores:
                        celula.fill = cores[str(celula.value)]
                    else:
                        celula.fill = branco

            # Salvar o arquivo Excel
            wb.save(arquivo_excel)

        except FileNotFoundError:
            print(f"O arquivo {arquivo_excel} não foi encontrado.")
        except Exception as e:
            print(f"Ocorreu um erro ao colorir as células: {e}")

    def salvar_dados_excel(status):
        global videoNumber
        dateCalculate()
        excelArchive = f'excel/{datetime.now().strftime("%d_%m_%Y")}.xlsx'
        # Tentar abrir o arquivo existente ou criar um novo se não existir
        try:
            wb = openpyxl.load_workbook(excelArchive)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active

        # Determinar a linha onde os dados serão inseridos
        linha_destino = videoNumber

        # Preparar os dados para inserção
        dados = [videoNumber, status, dateSelect]

        # Inserir dados na linha específica
        for col, valor in enumerate(dados, start=1):
            sheet.cell(row=linha_destino, column=col, value=valor)

        # Ajustar largura das colunas
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Salvar o arquivo Excel
        wb.save(excelArchive)

        logs.colorir_celulas(excelArchive)

# Categorias de video para upar
"""
//Futebol// 3
dls = Dream League Soccer
fcmob = FC Mobile
wsc = World Soccer Champs

//Caminhoes + Carros// 8
wtds = World Truck Driver Simulator
trd = The Road Driver
gto = Global Truck Online
farmings = Farming Simulator
cxdr = Carx Drift Racing 2
motow = Moto Wheelie 3D
drivers = Drivers Jobs
carp = Car Parking

//Infantil// 6
pkxd = Pk Xd
bloxf = Blox Fruits
roblox = Roblox
toca = Toca Life
sims = The Sims Freeplay
subway = Subway Surfers

//Outros// 5
gangstar = Gangstar Vegas
madout = Madout 2 Big City
sf2 = Shadow Fight 2
pvz = Plants vs Zombies 2
bit = Bit Life BR
"""
start = 2
end = 4
jumpDay = 10
firstDate = 10
postar = ["sims"]

for videoType in postar:
    info = conteudos[videoType]
    foundSelector = info["pasta"]
    tittlePath = info["titulo_arquivo"]
    descPath = info["desc_arquivo"]
    hour = info["horario"]

    for videoNumber in range(start, end+1):
        aboutVideoInfos()
        time_start = time.time()
        openNav("https://studio.youtube.com/channel/UCYUNcbRJKyGbmNjw2rSdh4A/videos/")
        time.sleep(5)
        maximo = clickImage(False, "images/windows/maximo.png", 0.8, 1)
        if maximo is False:
            minimo = clickImage(True, "images/windows/minimo.png", 0.85, 1)
        time.sleep(12)
        openConsole()
        steps.step1Upload()
        if errorInfo == True:
            continue
        steps.step2Edit()
        if errorInfo == True:
            continue
        
        logs.salvar_dados_excel('Posted')
        closeNav()
        firstDate += jumpDay
        timeCalc()
        ntfy(f"✅✅✅✅✅\nPrevisão de termino: {horario_estimado}\nContas Verficadas {videoNumber}/{end}\nCategoria: {videoType}\nVersão:{version}")
