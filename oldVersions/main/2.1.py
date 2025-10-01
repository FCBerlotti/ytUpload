import pyautogui as pg
import time
import pyperclip
import requests
import subprocess
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from conteudos import conteudos
import json
from pathlib import Path
import tkinter as tk
from tkinter import simpledialog
import inspect

base_path = Path(__file__).parent
consoleCounter = 0
contador = 0
time_soma = 0
videoUploaded = False
errorInfo = False
startDate = (date.today()).day
fullStartDate = date.today()
version = "2.1"

"Funcionalidades adicionadas na versão 2.0"
# Pagina de login com interface
# Corrigido bug de duplicação no erro "videoPosted 305"
# Agora a identificação da linha de erro é inteligente, capaz de identificar a exata linha em que o código está escrito

"Funcionalidades para adicionar"
# Correção na função do excel --escrever nome da funcao aqui-- ela anotava apenas o numero do video e se deu erro ou nao, agora ela anota a categoria e quais videos foram enviados ou não -- adicionar
# Organização - Funções agora são separadas por classes -- adicionar
# Melhoramento em "def timeCalc()" agora a função é capaz de calcular o tempo para envio de todas as categorias e não só da categoria atual -- adicionar
# Melhorar a semantica das variaveis e funções = nome das variaveis -- adicionar
# Criação das pastas gringas para envios gringos -- adicionar
# Criar funcao para chamar o ntfy no codigo, com todas as notificacoes possiveis -- adicionar
# Padronizar com o RDP, só rodar naquela resolucao -- adicionar
# Montar passo a passo para configurar um novo pc, com o RDP e sem o RDP -- adicionar
# Adicionar verificação para quando iniciar o bot, ele verificar todos os nomes dos conteudos antes de comecar a rodar --adicionar 
# Adicionar um campo de contagem de 24h no dashboard --adicionar
# Na dashboard, adicionar uma funcao para clicar e reenviar os videos com erros -- adicionar
# Padronizar a descricao, para cada tipo de linguagem, sendo possivel o usuario colocar apenas as tags e o link de download para cada conteudo na dashboard -- adicionar

def ntfy(message):
    try:
        if userSelect == "berlotti":
            requests.post(('https://ntfy.sh/ytUploadNotifier'), data=(message).encode(encoding='utf-8'))
        elif userSelect == "fabio":
            requests.post(('https://ntfy.sh/ytUploadNotifierfabio'), data=(message).encode(encoding='utf-8'))
    except requests.exceptions.HTTPError as errh:
        print("Http Error:", errh)
    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting:", errc)
    except requests.exceptions.Timeout as errt:
        print("Timeout Error:", errt)
    except requests.exceptions.RequestException as err:
        print("OOps: Something Else", err)

def timeCalc():
    global tempo_estimado, horario_estimado, time_soma, contador, time_final, time_start, end, videoNumber
    time_final = time.time()
    
    contador += 1
    time_soma += time_final-time_start
    tempo_estimado = (time.time() + ((time_soma / contador)*(end-videoNumber)))
    horario_estimado = datetime.fromtimestamp((time.time() + ((time_soma / contador)*(end-videoNumber)))).strftime('%Y-%m-%d %H:%M:%S')

class navigator:
    def openNav(site):
        caminho_chrome = "C:/Program Files/Google/Chrome/Application/chrome.exe"  # Atualize o caminho se necessário
        comando = [caminho_chrome, site]
        subprocess.Popen(comando)

    def openConsole():
        global consoleCounter, VopenConsole
        VopenConsole = "v27"
        pg.hotkey("ctrl", "shift", "i")
        for i in range(0,100):
            inspecActivity = pg.locateCenterOnScreen(f"users/{userSelect}/images/console/inspecActivity.png", confidence=0.95)
            if inspecActivity is not None and len(inspecActivity) == 2:
                pg.hotkey("ctrl", "p")
                break
            time.sleep(0.1)
        time.sleep(1)
        pg.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pg.write(">console")
        panelVerify = clickImage(False, f"users/{userSelect}/images/console/panel.png", 0.95, 25)
        if panelVerify is True:
            time.sleep(0.25)
            pg.press("enter")
            pyperclip.copy('allow pasting')
            time.sleep(1)
            pg.hotkey('ctrl', 'v')
            time.sleep(1)
            pg.press('enter')
            time.sleep(0.5)
            pg.write("allow pasting")
            time.sleep(1)
            pg.press("enter")
            time.sleep(0.75)
        else:
            pg.hotkey("ctrl", "shift", "i")
            if consoleCounter < 2:
                navigator.openConsole()
                consoleCounter += 1

    def closeNav(codeStep):
        pg.hotkey('ctrl', 'w')
        time.sleep(1)
        pg.press('enter')
        time.sleep(1)
        pg.press('esc')
        time.sleep(0.5)
        pg.press("enter")
        if codeStep == "Final":
            time.sleep(1)
            pg.hotkey('alt', 'f4')
            time.sleep(0.5)
            pg.press('esc')
    
    def openNavConfig():
        maximo = clickImage(False, f"users/{userSelect}/images/windows/maximo.png", 0.8, 1)
        if maximo is False:
            clickImage(True, f"users/{userSelect}/images/windows/minimo.png", 0.85, 1)
        for i in range(0, 120):
            studioOpen = clickImage(False, f'users/{userSelect}/images/youtube/studioOpen.png', 0.9, 1)
            if studioOpen:
                break
            time.sleep(0.1)
        navigator.openConsole()

def execJs(path):
    global desc, tittle
    try:
        # Abrir o arquivo JS e ler o conteúdo
        with open(path, "r", encoding="utf-8") as file:
            js_script = file.read()
            if 'tittleDescThumb' in path:
                js_script = f'var descricaoDesejada = {json.dumps(desc)}; var tituloDesejado = {json.dumps(tittle)};' + js_script

            # Copiar o conteúdo do script para a área de transferência
            pyperclip.copy(js_script)

            # Usar pg para colar o script no console
            pg.hotkey('ctrl', 'v')  # Cola o conteúdo copiado
            time.sleep(0.25)
            pg.press('enter') 
            pg.press('enter')
            pg.press('enter') # Pressiona Enter para executar o script
            time.sleep(1)

    except FileNotFoundError:
        print(f"O arquivo {path} não foi encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def clickImage(click, imageSelect, conf, attempts):
    for i in range(0, attempts):
        imageFind = pg.locateCenterOnScreen(imageSelect, confidence=conf)
        if imageFind is not None and len(imageFind) == 2:
            if click is True:
                pg.moveTo(imageFind[0], imageFind[1])
                time.sleep(0.5)
                pg.click()
                time.sleep(0.5)

            return True
        time.sleep(0.1)
    return False

def errorFunction(etapa):
    global errorInfo, videoNumber, firstDate, jumpDay
    errorInfo = True
    firstDate += jumpDay
    navigator.closeNav("")
    logs.salvar_dados_excel(f'Error {etapa}')
    errorList.append((videoType, videoNumber, dateSelect))
    timeCalc()
    ntfy(f"❌❌❌❌❌\nPrevisão de termino: {horario_estimado}\nNumero do Video: {videoNumber}/{end}\nCategoria: {videoType}\nEtapa do Erro: {etapa}\nVersão:{version}\nOpenConsole: {VopenConsole}")

def aboutVideoInfos():
    global tittle, desc, titlePath

    try:
        with open(f"{titlePath}", "r", encoding="utf-8") as file:
            tittle = file.read()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{titlePath}' não foi encontrado.")
    except Exception as e:
        print(f"Erro ao ler o arquivo: {str(e)}")
    
    try:
        with open(f"{descPath}", "r", encoding="utf-8") as arquivo:
            desc = arquivo.read()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{descPath}' não foi encontrado.")
    except Exception as e:
        print(f"Erro ao ler o arquivo: {str(e)}")

class calculateDates:    
    def dateCalculate():
        global dateSelect, firstDate
        today = datetime.now()
        dateSelect = (today + timedelta(days=firstDate)).strftime("%d/%m/%Y")

    def goalDateF(goalDateInput, startDateInput):
        global firstDate
        try:
            futureDate = datetime.strptime(goalDateInput, "%d/%m/%Y").date()
            calcDate = futureDate - startDateInput
            firstDate = calcDate.days
        except ValueError:
            print("Formato de data inválido. Tente novamente no formato DD/MM/AAAA.")

class videoConfigs:
    def tittleDescThumb():
        execJs('js/youtube/tittleDescThumb.js')

class steps:
    def step1Upload():
        global videoNumber, videoType, foundSelectorVideo
        time.sleep(5)
        execJs('js/youtube/upload.js')
        time.sleep(1)
        pg.hotkey('ctrl', 'shift', 'i')
        time.sleep(2)
        
        def foundSelectorF():
            uploadClick = clickImage(True, f'users/{userSelect}/images/youtube/upload.png', 0.9, 50)
            if uploadClick is False:
                errorFunction(f"uploadClick {inspect.currentframe().f_lineno}")
                return
            time.sleep(1)
            linkFolderClick = clickImage(True, f'users/{userSelect}/images/windows/linkFolder.png', 0.9, 20)
            if linkFolderClick is False:
                clickImage(True, f'users/{userSelect}/images/youtube/upload.png', 0.9, 50)
                time.sleep(2.5)
                linkFolderClick = clickImage(True, f'users/{userSelect}/images/windows/linkFolder.png', 0.9, 20)
            if linkFolderClick is True:
                pg.write(foundSelectorVideo)
                time.sleep(1)
                pg.press('enter')
                time.sleep(1.5)
                pg.hotkey('ctrl', 'f')
                time.sleep(1)
                pg.write(f'"{videoType}{videoNumber}.mp4"', 0.15)
                time.sleep(0.75)
                selectVideo = clickImage(True, f'users/{userSelect}/images/windows/mp4.png', 0.9, 80)
                if selectVideo is True:
                    time.sleep(0.5)
                    pg.press('enter')
                    time.sleep(0.5)
                    print("video selecionado e enviado")
            else:
                errorFunction(f"linkFolderClick {inspect.currentframe().f_lineno}")
                return
        foundSelectorF()
    
    def step2Edit():
        global videoUploaded, videoType, foundSelectorThumb
        print("iniciou step2")
        detalhes = clickImage(False, f"users/{userSelect}/images/youtube/detalhes.png", 0.9, 100)
        if detalhes is True:
            navigator.openConsole()
            time.sleep(1)
            videoConfigs.tittleDescThumb()
            time.sleep(2)
            linkFolderClick = clickImage(True, f'users/{userSelect}/images/windows/linkFolder.png', 0.9, 20)
            if linkFolderClick is True:
                pg.write(foundSelectorThumb)
                time.sleep(1)
                pg.press('enter')
                time.sleep(1.5)
                pg.hotkey('ctrl', 'f')
                time.sleep(1)
                pg.write(f'"{videoType}.png"')
                selectThumb = clickImage(True, f'users/{userSelect}/images/windows/thumbnail.png', 0.95, 80)
                if selectThumb is True:
                    time.sleep(0.5)
                    pg.press('enter')
                    time.sleep(0.5)
        else:
            errorFunction(f"LinkFolderClick {inspect.currentframe().f_lineno}")
            return
        
        time.sleep(5)
        pg.hotkey("ctrl", 'shift', "i")
        time.sleep(1)
        navigator.openConsole()
        time.sleep(1)
        execJs('js/youtube/next.js')
        time.sleep(1)

        copyright = clickImage(False, f'users/{userSelect}/images/youtube/copyright.png', 0.85, 1)
        copyright2 = clickImage(False, f'users/{userSelect}/images/youtube/copyright2.png', 0.85, 1)

        for ecodeverify in range(0, 200):
            if copyright is False and copyright2 is False:
                copyright = clickImage(False, f'users/{userSelect}/images/youtube/copyright.png', 0.85, 1)
                copyright2 = clickImage(False, f'users/{userSelect}/images/youtube/copyright2.png', 0.85, 1)
            else:
                break
            time.sleep(0.1)

        if copyright is True or copyright2 is True:
            time.sleep(1)
            execJs('js/youtube/next2.js')
            time.sleep(1)
            verifyPage = clickImage(False, f'users/{userSelect}/images/youtube/verifyPage.png', 0.9, 120)
            if verifyPage is True:
                execJs('js/youtube/date.js')
                time.sleep(5)
                dateChange = clickImage(True, f'users/{userSelect}/images/youtube/2025.png', 0.9, 40)
                if dateChange is True:
                    time.sleep(1.5)
                    pg.hotkey('ctrl', 'a')
                    time.sleep(0.5)
                    pg.write(f'{dateSelect}') 
                    time.sleep(1)
                    pg.press('enter')
                    time.sleep(1)
                    hourSelect = clickImage(True, f'users/{userSelect}/images/youtube/hourSelect.png', 0.9, 50)
                    time.sleep(0.75)
                    if hourSelect is True:
                        pg.hotkey('ctrl', 'a')
                        time.sleep(1)
                        pg.write(hour, 0.05)
                        time.sleep(0.5)
                        pg.press('enter')
                        time.sleep(1)
                    else:
                        errorFunction(f"hourSelect {inspect.currentframe().f_lineno}")
                        return
                    programar = clickImage(True, f'users/{userSelect}/images/youtube/programar.png', 0.9, 40)
                    if programar is False:
                        errorFunction(f"programar {inspect.currentframe().f_lineno}")
                    time.sleep(1)
            videoPosted = clickImage(False, f'users/{userSelect}/images/youtube/videoPosted.png', 0.88, 150) #TODO ALTERRAR REPETICOES DE TENTATIVAS DE ACORDO COM O USUARIO FABIO = 250
            if videoPosted is True or programar is True:
                videoUploaded = True
            else:
                videoUploaded = False
                errorFunction(f"videoPosted {inspect.currentframe().f_lineno}")
                return
        else:
            errorFunction(f"copyright {inspect.currentframe().f_lineno}")
            return
        
class logs:
    def colorir_celulas(arquivo_excel):
        try:
            # Carregar o arquivo Excel
            wb = openpyxl.load_workbook(arquivo_excel)
            sheet = wb.active

            # Definir cores para os textos
            verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            roxo = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
            branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Dicionário mapeando textos para cores
            cores = {
                "Error": roxo,
                "Posted": verde,
            }

            # Loop pelas células do intervalo especificado
            for linha in range(1, 2400 + 1): 
                for coluna in range(1, 3):  # Colunas de A até C(3)
                    celula = sheet.cell(row=linha, column=coluna)
                    # Verifica se o valor da célula existe e está no dicionário de cores
                    if celula.value and str(celula.value) in cores:
                        celula.fill = cores[str(celula.value)]
                    else:
                        celula.fill = branco

            # Salvar o arquivo Excel
            wb.save(arquivo_excel)

        except FileNotFoundError:
            print(f"O arquivo {arquivo_excel} não foi encontrado.")
        except Exception as e:
            print(f"Ocorreu um erro ao colorir as células: {e}")

    def salvar_dados_excel(status):
        global videoNumber
        excelArchive = f'excel/{datetime.now().strftime("%d_%m_%Y")}.xlsx'
        # Tentar abrir o arquivo existente ou criar um novo se não existir
        try:
            wb = openpyxl.load_workbook(excelArchive)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active

        # Determinar a linha onde os dados serão inseridos
        linha_destino = videoNumber

        # Preparar os dados para inserção
        dados = [videoNumber, status, dateSelect]

        # Inserir dados na linha específica
        for col, valor in enumerate(dados, start=1):
            sheet.cell(row=linha_destino, column=col, value=valor)

        # Ajustar largura das colunas
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Salvar o arquivo Excel
        wb.save(excelArchive)

        logs.colorir_celulas(excelArchive)

class users:
    def userSelectF():
        global userSelect, ytStudioLink
        if userSelect == "berlotti":
            ytStudioLink = "https://studio.youtube.com/channel/UCPDa_GVRpoAwRVCSnAZ8V_A/videos/" # TODO mudar para txt inteligente depois
        elif userSelect == "fabio":
            ytStudioLink = "https://studio.youtube.com/channel/UC3DZUNHs1SsdA8YE2PbcSZg"

    def ler_configuracao():
        config_path = base_path / "configs/loginConfig.txt"
        
        configuracoes = {}
        try:
            with open(config_path, "r") as f:
                for linha in f:
                    # Remove espaços em branco e quebras de linha
                    linha = linha.strip()
                    # Ignora linhas vazias
                    if not linha:
                        continue
                    
                    # Divide a linha no primeiro '=' para separar a chave e o valor
                    chave, valor = linha.split("=", 1)
                    
                    # Salva no dicionário de configurações
                    configuracoes[chave] = valor
            
            print(f"Configurações lidas com sucesso do arquivo: {config_path}")
            return configuracoes
            
        except FileNotFoundError:
            print(f"Erro: O arquivo de configuração {config_path} não foi encontrado.")
            return None
        except Exception as e:
            print(f"Ocorreu um erro ao ler o arquivo de configuração: {e}")
            return None

pg.hotkey("win", "m")
config = users.ler_configuracao()
userSelect = config.get("username")

def dadosIniciais():
    global start, end, jumpDay, errorList, postar, foundSelectorVideo, foundSelectorThumb, userSelect, attemptsWhile, contentLanguague, ytStudioLink, goalDate
    foundSelectorVideo = f"{base_path / 'videos'}"
    foundSelectorThumb = f"{base_path / 'thumbs'}"
    start = 15
    end = 22
    jumpDay = 10
    goalDate = "15/10/2025"
    userLanguague = "NULL" # disponibilizar linguagens para a preferencia do usuario
    contentLanguague = "NULL" # pt-br / en-us / es-es # Usar para escolher qual tipo de conteudo vai ser postado e em qual linguagem vai ser postado
    attemptsWhile = 0
    users.userSelectF()
    calculateDates.goalDateF(goalDate, fullStartDate)
 
   #""" TODO criar verificacao para quando atingir o limite de envio diario
    #js path do erro do yt de maximo de envios : document.querySelector("#dialog > div > ytcp-animatable.button-area.metadata-fade-in-section.style-scope.ytcp-uploads-dialog > div > div.left-button-area.style-scope.ytcp-uploads-dialog > ytcp-ve > div.error-short.style-scope.ytcp-uploads-dialog").textContent
    #retorna isso se tiver textContent no final: 'O limite diário de envios foi alcançado'
    #colocar time sleep para enviar isso dps que der o next     
    #o ultimo codigo do gemini funciona se colar e logo em seguida dar ctrl shift i, a pagina precisa estar clicada para copiar o texto"""
    
errorList = [] #("sf2", "11", "01/08/2026") item para teste
postar = ["footl", "hungry", "boxmob"]

for videoType in postar:
    dadosIniciais()
    info = conteudos[videoType]
    titlePath = info["titulo_arquivo"]
    descPath = f"users/{userSelect}/{info['desc_arquivo']}"
    hour = info["horario"]
    
    actualDate = (date.today()).day 
    if actualDate > startDate:
        firstDate -= 1
    if videoType == "hungry" or videoType == "boxmob":
        calculateDates.goalDateF("25/09/2025", fullStartDate)
        start = 13
        end = 22

    if videoType == "TEMPLATE":
        print("")
    else:
        print("TEMPLATE")
        #calculateDates.goalDateF("17/08/2025", fullStartDate)
        
    ntfy(f"INICIANDO ENVIO DE {videoType} {int(postar.index(videoType))+1}/{len(postar)}\nVersão: {version}\nUser: {userSelect}")

    for videoNumber in range(start, end+1):
        time_start = time.time()
        while videoNumber > 10:
            videoNumber -= 10
        aboutVideoInfos()
        calculateDates.dateCalculate()
        errorInfo = False
        navigator.openNav(ytStudioLink)
        time.sleep(5)
        navigator.openNavConfig()
        steps.step1Upload()
        if errorInfo == True:
            continue
        steps.step2Edit()
        if errorInfo == True:
            continue
        logs.salvar_dados_excel('Posted')
        navigator.closeNav("")
        firstDate += jumpDay
        timeCalc()
        ntfy(f"✅✅✅✅✅\nPrevisão de termino: {horario_estimado}\nVideos: {videoNumber}/{end}\nCategoria: {videoType}\nVersão:{version}\nOpenConsole: {VopenConsole}\nUser: {userSelect}")

ntfy(f"INICIANDO SESSAO DE ERROS\nCategoria: {videoType}\nNumero do Video: {videoNumber}\nData da Postagem: {dateSelect}\nVersão: {version}\nOpenConsole: {VopenConsole}\nUser: {userSelect}")
while len(errorList) > 0 or attemptsWhile < 5:
    attemptsWhile += 1
    for (videoType, videoNumber, dateSelect) in errorList:
        errorList.remove((videoType, videoNumber, dateSelect))
        errorInfo = False
        info = conteudos[videoType]
        titlePath = info["titulo_arquivo"]
        descPath = info["desc_arquivo"]
        hour = info["horario"]
        dateSelect = dateSelect

        time_start = time.time()
        aboutVideoInfos()
        navigator.openNav(ytStudioLink)
        time.sleep(5)
        navigator.openNavConfig()
        steps.step1Upload()
        if errorInfo == True:
            continue
        steps.step2Edit()
        if errorInfo == True:
            continue
        navigator.closeNav("")
        timeCalc()
        ntfy(f"✅✅✅✅✅Error Function\nPrevisão de termino: {horario_estimado}\nNumero do Video: {videoNumber}/{end}\nCategoria: {videoType}\nVersão: {version}\nOpenConsole: {VopenConsole}\nUser: {userSelect}")

navigator.closeNav("Final")
ntfy("TODOS OS ENVIOS FORAM FINALIZADOS")

"""ideia para a versao 2.0, fazer um menu e um campo para inserir um novo tipo de conteudo, nesse campo ser possivel colocar
descricao padrao, titulo padrao, ou entao apenas substituir o link padrao de alguma pagina, mais p frente, ser possivel implementar sistema com o YT
fazendo com que seja possivel alterar todas as descricoes dos videos, exemplo: quero mudar a descricao ou o link de algum conteudo."""