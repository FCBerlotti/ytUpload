import pyautogui as pg
import time
import pyperclip
import requests
import subprocess
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from conteudos import conteudos
import json
from pathlib import Path

base_path = Path(__file__).parent

consoleCounter = 0
contador = 0
time_soma = 0
videoUploaded = False
errorInfo = False
startDate = (date.today()).day 
version = "1.5"

"Funcionalidades adicionadas na versão 1.5"
# Reformulação na abertura e fechamento do navegador "def closeNav()" -- adicionado
# Correção de erro ao fechar navegador, quando envio dava erro, não fechava a pagina do envio -- adicionado
# Variavel para atualizacao de data, caso comece em um dia e termine em outro
# Caminhos para as pastas agora são definidos automaticos, de acordo com o computador em que está, utilizando o 'base_path'
# Correção na ativação do allow pasting em 'openConsole()' -- adicionado

"Funcionalidades para adicionar"
# Correção na função do excel --escrever nome da funcao aqui-- ela anotava apenas o numero do video e se deu erro ou nao, agora ela anota a categoria e quais videos foram enviados ou não -- adicionar
# Organização - Funções agora são separadas por classes -- adicionar
# Melhoramento em "def timeCalc()" agora a função é capaz de calcular o tempo para envio de todas as categorias e não só da categoria atual -- adicionar
# Melhorar a semantica das variaveis e funções = nome das variaveis -- adicionar
# Criação das pastas gringas para envios gringos

def ntfy(message):
    try:
        requests.post(('https://ntfy.sh/ytUploadNotifier'), data=(message).encode(encoding='utf-8'))
    except requests.exceptions.HTTPError as errh:
        print("Http Error:", errh)
    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting:", errc)
    except requests.exceptions.Timeout as errt:
        print("Timeout Error:", errt)
    except requests.exceptions.RequestException as err:
        print("OOps: Something Else", err)

def timeCalc():
    global tempo_estimado, horario_estimado, time_soma, contador, time_final, time_start, end, videoNumber
    time_final = time.time()
    
    contador += 1
    time_soma += time_final-time_start
    tempo_estimado = (time.time() + ((time_soma / contador)*(end-videoNumber)))
    horario_estimado = datetime.fromtimestamp((time.time() + ((time_soma / contador)*(end-videoNumber)))).strftime('%Y-%m-%d %H:%M:%S')

def openNav(site):
    caminho_chrome = "C:/Program Files/Google/Chrome/Application/chrome.exe"  # Atualize o caminho se necessário
    comando = [caminho_chrome, site]
    subprocess.Popen(comando)

def openConsole():
    global consoleCounter
    pg.hotkey("ctrl", "shift", "i")
    for i in range(0,50):
        inspecActivity = pg.locateCenterOnScreen("images/console/inspecActivity.png", confidence=0.95)
        if inspecActivity is not None and len(inspecActivity) == 2:
            pg.hotkey("ctrl", "p")
            break
        time.sleep(0.1)
    time.sleep(1)
    pg.hotkey('ctrl', 'a')
    time.sleep(0.5)
    pg.write(">console")
    time.sleep(0.5)
    panelVerify = pg.locateCenterOnScreen("images/console/panel.png", confidence=0.95)
    if panelVerify is not None and len(panelVerify) == 2:
        time.sleep(0.25)
        pg.press("enter")
        pyperclip.copy('allow pasting')
        time.sleep(1)
        pg.hotkey('ctrl', 'v')
        time.sleep(0.25)
        pg.press('enter')
        time.sleep(0.2)
        pg.write("allow pasting")
        time.sleep(0.25)
        pg.press("enter")
        time.sleep(0.75)
        pg.press('up')
        time.sleep(0.5)
        pg.press('enter')
    else:
        pg.hotkey("ctrl", "shift", "i")
        if consoleCounter < 2:
            openConsole()
            consoleCounter += 1

def execJs(path):
    global desc, tittle
    try:
        # Abrir o arquivo JS e ler o conteúdo
        with open(path, "r", encoding="utf-8") as file:
            js_script = file.read()
            if 'tittleDescThumb' in path:
                js_script = f'var descricaoDesejada = {json.dumps(desc)}; var tituloDesejado = {json.dumps(tittle)};' + js_script

            # Copiar o conteúdo do script para a área de transferência
            pyperclip.copy(js_script)

            # Usar pg para colar o script no console
            pg.hotkey('ctrl', 'v')  # Cola o conteúdo copiado
            time.sleep(0.25)
            pg.press('enter') 
            pg.press('enter')
            pg.press('enter') # Pressiona Enter para executar o script
            time.sleep(1)

    except FileNotFoundError:
        print(f"O arquivo {path} não foi encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def clickImage(click, imageSelect, conf, attempts):
    for i in range(0, attempts):
        imageFind = pg.locateCenterOnScreen(imageSelect, confidence=conf)
        if imageFind is not None and len(imageFind) == 2:
            if click is True:
                pg.moveTo(imageFind[0], imageFind[1])
                time.sleep(0.5)
                pg.click()
                time.sleep(0.5)

            return True
        time.sleep(0.1)
    return False

def errorFunction(etapa):
    global errorInfo, videoNumber, firstDate, jumpDay
    errorInfo = True
    firstDate += jumpDay
    closeNav("")
    logs.salvar_dados_excel('Error')
    errorList.append((videoType, videoNumber, dateSelect))
    timeCalc()
    ntfy(f"❌❌❌❌❌\nPrevisão de termino: {horario_estimado}\nNumero do Video: {videoNumber}/{end}\nCategoria: {videoType}\nEtapa do Erro: {etapa}\nVersão:{version}")

def closeNav(codeStep):
    pg.hotkey('ctrl', 'w')
    time.sleep(1)
    pg.press('enter')
    time.sleep(1)
    pg.press('esc')
    time.sleep(0.5)
    pg.press("enter")
    if codeStep == "Final":
        time.sleep(1)
        pg.hotkey('alt', 'f4')
        time.sleep(0.5)
        pg.press('esc')

def aboutVideoInfos():
    global tittle, desc, titlePath

    try:
        with open(f"{titlePath}", "r", encoding="utf-8") as file:
            tittle = file.read()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{titlePath}' não foi encontrado.")
    except Exception as e:
        print(f"Erro ao ler o arquivo: {str(e)}")
    
    try:
        with open(f"{descPath}", "r", encoding="utf-8") as arquivo:
            desc = arquivo.read()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{descPath}' não foi encontrado.")
    except Exception as e:
        print(f"Erro ao ler o arquivo: {str(e)}")

class calculateDates:    
    def dateCalculate():
        global dateSelect, firstDate
        today = datetime.now()
        dateSelect = (today + timedelta(days=firstDate)).strftime("%d/%m/%Y")
    def goalDate():
        hoje = date.today()
        alvo = "2025-08-27"
        try:
            ano, mes, dia = map(int, alvo.split('-'))
            data_futura = date(ano, mes, dia)
            calculoDate = data_futura - hoje
        except ValueError:
            print("Formato de data inválido. Tente novamente no formato AAAA-MM-DD.")

class videoConfigs:
    def tittleDescThumb():
        execJs('js/youtube/tittleDescThumb.js')

class steps:
    def step1Upload():
        global videoNumber, videoType, foundSelectorVideo
        time.sleep(5)
        execJs('js/youtube/upload.js')
        time.sleep(1)
        pg.hotkey('ctrl', 'shift', 'i')
        time.sleep(2)
        
        def foundSelectorF():
            uploadClick = clickImage(True, 'images/youtube/upload.png', 0.9, 50)
            if uploadClick is False:
                errorFunction("uploadClick 193")
                return
            time.sleep(1)
            linkFolderClick = clickImage(True, 'images/windows/linkFolder.png', 0.9, 20)
            if linkFolderClick is False:
                clickImage(True, 'images/youtube/upload.png', 0.9, 50)
                time.sleep(2.5)
                linkFolderClick = clickImage(True, 'images/windows/linkFolder.png', 0.9, 20)
            if linkFolderClick is True:
                pg.write(foundSelectorVideo)
                time.sleep(1)
                pg.press('enter')
                time.sleep(1.5)
                pg.hotkey('ctrl', 'f')
                time.sleep(1)
                pg.write(f'"{videoType}{videoNumber}.mp4"', 0.15)
                time.sleep(0.75)
                selectVideo = clickImage(True, 'images/windows/mp4.png', 0.9, 80)
                if selectVideo is True:
                    time.sleep(0.5)
                    pg.press('enter')
                    time.sleep(0.5)
                    print("video selecionado e enviado")
            else:
                errorFunction("linkFolderClick 217")
                return
        foundSelectorF()
    
    def step2Edit():
        global videoUploaded, videoType, foundSelectorThumb
        print("iniciou step2")
        detalhes = clickImage(False, "images/youtube/detalhes.png", 0.9, 100)
        if detalhes is True:
            openConsole()
            time.sleep(1)
            videoConfigs.tittleDescThumb()
            time.sleep(2)
            linkFolderClick = clickImage(True, 'images/windows/linkFolder.png', 0.9, 20)
            if linkFolderClick is True:
                pg.write(foundSelectorThumb)
                time.sleep(1)
                pg.press('enter')
                time.sleep(1.5)
                pg.hotkey('ctrl', 'f')
                time.sleep(1)
                pg.write(f'"{videoType}.png"')
                selectThumb = clickImage(True, 'images/windows/thumbnail.png', 0.9, 80)
                if selectThumb is True:
                    time.sleep(0.5)
                    pg.press('enter')
                    time.sleep(0.5)
        else:
            errorFunction("LinkFolderClick 246")
            return
        
        time.sleep(5)
        pg.hotkey("ctrl", 'shift', "i")
        time.sleep(1)
        openConsole()
        time.sleep(1)
        execJs('js/youtube/next.js')
        time.sleep(1)

        copyright = clickImage(False, 'images/youtube/copyright.png', 0.85, 1)
        copyright2 = clickImage(False, 'images/youtube/copyright2.png', 0.85, 1)

        for i in range(0, 300):
            if copyright is False and copyright2 is False:
                copyright = clickImage(False, 'images/youtube/copyright.png', 0.85, 1)
                copyright2 = clickImage(False, 'images/youtube/copyright2.png', 0.85, 1)
            else:
                break
            time.sleep(0.1)

        if copyright is True or copyright2 is True:
            time.sleep(1)
            execJs('js/youtube/next2.js')
            time.sleep(1)
            verifyPage = clickImage(False, 'images/youtube/verifyPage.png', 0.9, 120)
            if verifyPage is True:
                execJs('js/youtube/date.js')
                time.sleep(5)
                dateChange = clickImage(True, 'images/youtube/2025.png', 0.9, 40)
                if dateChange is True:
                    time.sleep(1.5)
                    pg.hotkey('ctrl', 'a')
                    time.sleep(0.5)
                    pg.write(f'{dateSelect}') 
                    time.sleep(1)
                    pg.press('enter')
                    time.sleep(1)
                    hourSelect = clickImage(True, 'images/youtube/hourSelect.png', 0.9, 50)
                    time.sleep(0.75)
                    if hourSelect is True:
                        pg.hotkey('ctrl', 'a')
                        time.sleep(1)
                        pg.write(hour, 0.05)
                        time.sleep(0.5)
                        pg.press('enter')
                        time.sleep(1)
                    else:
                        errorFunction("hourSelect 296")
                        return
                    clickImage(True, 'images/youtube/programar.png', 0.9, 40)
                    time.sleep(1)
            videoPosted = clickImage(False, 'images/youtube/videoPosted.png', 0.88, 150)
            if videoPosted is True:
                videoUploaded = True
            else:
                videoUploaded = False
                errorFunction("videoPosted 305")
                return
        else:
            errorFunction("copyright 308")
            return
        
class logs:
    def colorir_celulas(arquivo_excel):
        try:
            # Carregar o arquivo Excel
            wb = openpyxl.load_workbook(arquivo_excel)
            sheet = wb.active

            # Definir cores para os textos
            verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            roxo = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
            branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Dicionário mapeando textos para cores
            cores = {
                "Error": roxo,
                "Posted": verde,
            }

            # Loop pelas células do intervalo especificado
            for linha in range(1, 2400 + 1): 
                for coluna in range(1, 3):  # Colunas de A até C(3)
                    celula = sheet.cell(row=linha, column=coluna)
                    # Verifica se o valor da célula existe e está no dicionário de cores
                    if celula.value and str(celula.value) in cores:
                        celula.fill = cores[str(celula.value)]
                    else:
                        celula.fill = branco

            # Salvar o arquivo Excel
            wb.save(arquivo_excel)

        except FileNotFoundError:
            print(f"O arquivo {arquivo_excel} não foi encontrado.")
        except Exception as e:
            print(f"Ocorreu um erro ao colorir as células: {e}")

    def salvar_dados_excel(status):
        global videoNumber
        excelArchive = f'excel/{datetime.now().strftime("%d_%m_%Y")}.xlsx'
        # Tentar abrir o arquivo existente ou criar um novo se não existir
        try:
            wb = openpyxl.load_workbook(excelArchive)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active

        # Determinar a linha onde os dados serão inseridos
        linha_destino = videoNumber

        # Preparar os dados para inserção
        dados = [videoNumber, status, dateSelect]

        # Inserir dados na linha específica
        for col, valor in enumerate(dados, start=1):
            sheet.cell(row=linha_destino, column=col, value=valor)

        # Ajustar largura das colunas
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Salvar o arquivo Excel
        wb.save(excelArchive)

        logs.colorir_celulas(excelArchive)

def openNavConfig():
    maximo = clickImage(False, "images/windows/maximo.png", 0.8, 1)
    if maximo is False:
        clickImage(True, "images/windows/minimo.png", 0.85, 1)
    for i in range(0, 120):
        studioOpen = clickImage(False, 'images/youtube/studioOpen.png', 0.9, 1)
        if studioOpen:
            break
        time.sleep(0.1)
    openConsole()

def dadosIniciais():
    global start, end, jumpDay, firstDate, errorList, postar, foundSelectorVideo, foundSelectorThumb, userSelect, attemptsWhile, contentLanguague
    foundSelectorVideo = f"{base_path / 'videos'}"
    foundSelectorThumb = f"{base_path / 'thumbs'}"
    start = 8
    end = 10
    jumpDay = 10
    firstDate = 0
    userSelect = "berlotti" # berlotti / fabio
    contentLanguague = "NULL" # pt-br / en-us / es-es # Usar para escolher qual tipo de conteudo vai ser postado e em qual linguagem vai ser postado
    attemptsWhile = 0

   #""" TODO criar verificacao para quando atingir o limite de envio diario
    #js path do erro do yt de maximo de envios : document.querySelector("#dialog > div > ytcp-animatable.button-area.metadata-fade-in-section.style-scope.ytcp-uploads-dialog > div > div.left-button-area.style-scope.ytcp-uploads-dialog > ytcp-ve > div.error-short.style-scope.ytcp-uploads-dialog").textContent
    #retorna isso se tiver textContent no final: 'O limite diário de envios foi alcançado'
    #colocar time sleep para enviar isso dps que der o next     
    #o ultimo codigo do gemini funciona se colar e logo em seguida dar ctrl shift i, a pagina precisa estar clicada para copiar o texto"""
    
errorList = []
postar = ["mine", "af", "monop", "wbus", "extreme", "tr2", "car2"]

for videoType in postar:
    dadosIniciais()
    info = conteudos[videoType]
    titlePath = info["titulo_arquivo"]
    descPath = info["desc_arquivo"]
    hour = info["horario"]

    actualDate = (date.today()).day 
    if actualDate > startDate:
        firstDate -= 1

    if videoType == "wbus" or videoType == "extreme":
        firstDate = 1
    elif videoType == "tr2" or videoType == "car2":
        firstDate = 2
    else:
        firstDate = 1
        start = 10
        
    ntfy(f"INICIANDO ENVIO DE {videoType} {int(postar.index(videoType))+1}/{len(postar)}\nVersão:{version}")

    for videoNumber in range(start, end+1):
        time_start = time.time()
        while videoNumber > 10:
            videoNumber -= 10
        aboutVideoInfos()
        calculateDates.dateCalculate()
        errorInfo = False
        openNav("https://studio.youtube.com/channel/UCPDa_GVRpoAwRVCSnAZ8V_A/videos/")
        time.sleep(5)
        openNavConfig()
        steps.step1Upload()
        if errorInfo == True:
            continue
        steps.step2Edit()
        if errorInfo == True:
            continue
        
        logs.salvar_dados_excel('Posted')
        closeNav("")
        firstDate += jumpDay
        timeCalc()
        ntfy(f"✅✅✅✅✅\nPrevisão de termino: {horario_estimado}\nVideos: {videoNumber}/{end}\nCategoria: {videoType}\nVersão:{version}")

while len(errorList) > 0 or attemptsWhile < 3:
    attemptsWhile += 1
    for videoType, videoNumber, dateSelect in errorList:
        errorInfo = False
        info = conteudos[videoType]
        titlePath = info["titulo_arquivo"]
        descPath = info["desc_arquivo"]
        hour = info["horario"]
        dateSelect = dateSelect

        time_start = time.time()
        ntfy(f"INICIANDO SESSAO DE ERROS\nCategoria: {videoType}\nNumero do Video: {videoNumber}\nData da Postagem: {dateSelect}\nVersão:{version}")
        aboutVideoInfos()
        openNav("https://studio.youtube.com/channel/UCPDa_GVRpoAwRVCSnAZ8V_A/videos/")
        time.sleep(5)
        openNavConfig()
        steps.step1Upload()
        if errorInfo == True:
            continue
        steps.step2Edit()
        if errorInfo == True:
            continue
        
        closeNav("")
        errorList.remove((videoType, videoNumber, dateSelect))
        timeCalc()
        ntfy(f"✅✅✅✅✅\nPrevisão de termino: {horario_estimado}\nNumero do Video: {videoNumber}/{end}\nCategoria: {videoType}\nVersão:{version}")

closeNav("Final")
ntfy("TODOS OS ENVIOS FORAM FINALIZADOS")

"""ideia para a versao 2.0, fazer um menu e um campo para inserir um novo tipo de conteudo, nesse campo ser possivel colocar
descricao padrao, titulo padrao, ou entao apenas substituir o link padrao de alguma pagina, mais p frente, ser possivel implementar sistema com o YT
fazendo com que seja possivel alterar todas as descricoes dos videos, exemplo: quero mudar a descricao ou o link de algum conteudo."""
